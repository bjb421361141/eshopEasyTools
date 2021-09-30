# -*- coding: utf-8 -*-

import os
import re
import shutil

import requests
import json

from PIL import Image
from openpyxl import load_workbook
from pyquery import PyQuery as pq

from commUtil.propertiesUtil import Properties
from login.alibabaLogin import AlibabaLogin

URL = "https://detail.1688.com/offer/553178615733.html\n" \
      "https://detail.1688.com/offer/655819846969.html\n" \
      "https://detail.1688.com/offer/654970638934.html\n" \
      "https://detail.1688.com/offer/623629351705.html\n" \
      "https://detail.1688.com/offer/628440305353.html\n" \
      "https://detail.1688.com/offer/623350174922.html\n" \
      "https://detail.1688.com/offer/649413088953.html"

# 静态参数
URL_LINK = "url_link"
MAIN_PIC = "mainPic"
DESC_PIC = "descPic"
STOCK_INF = "stockInf"

COOKIES = "cna=iasVGVu3nj4CAbf7XYGq6VK5; _bl_uid=C2kbqobOe0deaR4gd0nzbjm838vI; UM_distinctid=1794725216cc7-07371a386069e6-30614f07-144000-1794725216d88a; taklid=d09ef3fcd8fb472b9f4bb69343e4fef0; ali_ab=183.251.93.129.1620396357714.7; lid=tb69481656; ali_apache_track=c_mid=b2b-1888131324|c_lid=tb69481656|c_ms=1; alicnweb=touch_tb_at=1624044364846|lastlogonid=tb69481656; xlly_s=1; _m_h5_tk=79d41008026287263e7b044a0fb8b6c2_1627315531261; _m_h5_tk_enc=da51ea45c6f0bbce7633167abb435679; cookie2=1a79067f96ba13b99c420d0576f437a4; t=5d630d2b98b2759cf58acc4dd530043d; _tb_token_=e1b80e537fdfa; __cn_logon__=false; CNZZDATA1253659577=1144812740-1620391771-https%3A%2F%2Fp4psearch.1688.com%2F|1627306510; _csrf_token=1627308137213; JSESSIONID=4F864C5A44EE1BB06967D5E2DE1D2549; tfstk=cRAfB9j71oqf31HagEgzQ0x4QTSNaBo5rxsvlpjGy-jK7J8F2sfYTM9rXdquw9Q5.; l=eBO8ZMUPjpyqRGsLBO5Bourza77TqIOb4uPzaNbMiInca6shxFiV7NCB82vyIdtjgt1e4etyG0JegdLHR3xBAxDDBkhy80ornxf..; isg=BImJwTXfDcaWIfFZ3sL5LMu0mLXj1n0IWy0cLSv_TnCvcqmEcyWq2HeqsNZEThVA"
properties_map = Properties("../login/config.properties").getproperties()

s = requests.Session()

def wrap_web_header():
    """
        可以使用simulation进行模拟登入 封装请求头信息
    :return: 请求头信息
    """
    global COOKIES
    global LOGIN_TOOL
    accept = r"text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9"
    user_agent = r"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.150 Safari/537.36"
    if COOKIES is None:
        LOGIN_TOOL = AlibabaLogin(properties_map.get("Alibaba").get("username"),
                                  properties_map.get("Alibaba").get("password"))
        LOGIN_TOOL.login()
        cookies_map = LOGIN_TOOL.get_cookies()
        if cookies_map is not None and len(cookies_map) > 0:
            COOKIES = ''
            for tmp_obj in cookies_map:
                if tmp_obj["name"] is not None and tmp_obj["value"] is not None:
                    COOKIES += tmp_obj["name"] + "=" + tmp_obj["value"]
        else:
            raise Exception("缺少cookies信息")  # 抛出异常信息

    headers = {
        "accept": accept,
        "accept-encoding": "gzip, deflate, br",
        "accept-language": "zh-CN,zh;q=0.9",
        "cache-control": "max-age=0",
        "cookie": COOKIES,
        "sec-fetch-dest": "document",
        "sec-fetch-mode": "navigate",
        "sec-fetch-site": "none",
        "sec-fetch-user": "?1",
        "upgrade-insecure-requests": "1",
        "user-agent": user_agent
    }
    return headers


def dir_download_images(a_url, a_filename, a_save_dir="C:\\Users\\Baijb\\Desktop\\产品模板\\"):
    """
        直接下载文件信息
    :param a_url:  文件路径
    :param a_filename:
    :param a_save_dir:
    :return:
    """
    try:
        imgs_data = requests.get(a_url, headers=wrap_web_header())
        save_name = str(a_save_dir + a_filename)
        if not os.path.exists(a_save_dir):
            os.makedirs(a_save_dir)
        with open(save_name, 'wb') as f:
            f.write(imgs_data.content)
    except OSError as err:
        print(err)


def resize_pic(pic_path, width, height, rtocut_flag=False):
    """
    缩放图片
    :param pic_path: 图片路径
    :param width: 宽
    :param height: 高
    :param rtocut_flag:按比例裁剪
    :return:
    """
    image = Image.open(pic_path)
    if image.mode == 'P':
        image = image.convert('RGB')
    base_width = width
    h_size = height
    if rtocut_flag:
        w_percent = base_width / float(image.size[0])
        h_size = int(float(image.size[1]) * float(w_percent))

    # 默认情况下，PIL使用Image.NEAREST过滤器进行大小调整，从而获得良好的性能，但质量很差。
    im_resized = image.resize((base_width, h_size), Image.ANTIALIAS)
    im_resized.save(pic_path)


def blend_brand_images(pic_base_path, pic_name, brand_pic_path="brand_pic.png"):
    """
    合并商标
    :return:
    """
    source_pic = Image.open(pic_base_path + pic_name)
    brand_pic = Image.open(brand_pic_path)
    if source_pic.size[0] != brand_pic.size[0] or source_pic.size[1] != brand_pic.size[1]:
        return
    source_pic = source_pic.convert('RGBA')
    final2 = Image.new("RGBA", source_pic.size)
    final2 = Image.alpha_composite(final2, source_pic)
    final2 = Image.alpha_composite(final2, brand_pic)
    # final2.show()
    final2 = final2.convert('RGB')
    final2.save(pic_base_path + pic_name.split(".")[0] + "_brand.jpg")


# def cv_resize

def grab_webpage(url):
    """
        获取指定的1688路径下的网页信息
    :param url:
    :return:  字典类 网页内容
    """
    webpage_info = {}
    resp_data = requests.get(url, headers=wrap_web_header()).content
    webpage_info[URL_LINK] = url
    doc = pq(resp_data)
    pic_objs = doc(".tab-trigger").items()
    main_pic_urls = []
    for pic in pic_objs:
        if pic.attr("data-imgs") is not None:
            json_obj = json.loads(pic.attr("data-imgs"))
            pic_url = json_obj["original"]
            pic_name = pic_url[json_obj["original"].rindex("/") + 1:]
            main_pic_urls.append({"pic_url": pic_url, "pic_name": pic_name})
    webpage_info[MAIN_PIC] = main_pic_urls

    # 详情信息是直接请求的
    desc_data_url = doc(".desc-lazyload-container").items()
    desc_pic_urls = []
    for tmp in desc_data_url:
        data_tfs_url = tmp.attr("data-tfs-url")
        desc_data_info = requests.get(data_tfs_url, headers=wrap_web_header()).content  # .decode('utf-8')
        img_lst = pq(desc_data_info)("img")
        for img_idx in range(0, img_lst.length):
            attr_src = pq(img_lst[img_idx]).attr('src')
            pic_url = pq(img_lst[img_idx]).attr('src')[attr_src.index("h"):attr_src.rindex("\\")]
            pic_name = pic_url[pic_url.rindex("/") + 1:]
            desc_pic_urls.append({"pic_url": pic_url, "pic_name": pic_name})
    webpage_info[DESC_PIC] = desc_pic_urls

    # 获取库存（尺码，价格信息）
    stock_inf_tr = doc(".obj-sku table tr")
    stock_inf = []
    for tr in stock_inf_tr:
        name = pq(tr).find("td.name").text()
        price = pq(tr).find("td.price em.value").text()
        count = pq(tr).find("td.count").text()
        tmp = {"name": name, "price": price, "count": count, "weight": "0"}  # weight 克重暂时使用0来进行测算
        stock_inf.append(tmp)
    webpage_info[STOCK_INF] = stock_inf
    return webpage_info


# Press the green button in the gutter to run the script.
def handel_pic_info(page_info, base_path="C:\\Users\\Baijb\\Desktop\\产品模板\\"):
    """
    处理图片信息
    :param page_info: 图片名称及下载路径
    :param base_path: 文件保存路径
    :return:
    """
    if not os.path.exists(base_path):
        try:
            os.makedirs(base_path + "01主图\\")
            os.makedirs(base_path + "02详情\\")
        except OSError as e:
            raise
    main_pic_inf = page_info[MAIN_PIC]
    main_pic_basepath = base_path + "01主图\\"
    for pic in main_pic_inf:
        try:
            dir_download_images(pic["pic_url"], pic["pic_name"], main_pic_basepath)
            if os.path.exists(main_pic_basepath + pic["pic_name"]):
                resize_pic(main_pic_basepath + pic["pic_name"], 800, 800)
                blend_brand_images(main_pic_basepath, pic["pic_name"])
        except BaseException as e:
            print(e)

    desc_pic_inf = page_info[DESC_PIC]
    desc_pic_basepath = base_path + "02详情\\"
    for pic in desc_pic_inf:
        dir_download_images(pic["pic_url"], pic["pic_name"], desc_pic_basepath)
        if os.path.exists(desc_pic_basepath + pic["pic_name"]):
            resize_pic(desc_pic_basepath + pic["pic_name"], 800, 800, True)

    return


def handel_stock_inf(stock_inf_lst, url_link, base_path="C:\\Users\\Baijb\\Desktop\\产品模板\\"):
    """
    复制成本信息模板，向指定位置设置库存信息并计算售卖价
        公式：
    :param stock_inf_lst: 库存信息列表
    :param base_path:文件保存路径
    :return:
    """
    shutil.copyfile("template.xlsx", base_path + "03成本计算.xlsx")
    wb = load_workbook(base_path + "03成本计算.xlsx")
    ws = wb.get_sheet_by_name('价格')
    # 从L13开始设置库存信息
    idx = 13
    for tk_inf in stock_inf_lst:
        # tmp = {"name": name, "price": price, "count": count, "weight": "0"}  # weight 克重暂时使用0来进行测算
        ws['L%s' % idx] = tk_inf["name"]
        ws['M%s' % idx] = tk_inf["count"]
        ws['N%s' % idx] = tk_inf["price"]
        ws['O%s' % idx] = tk_inf["weight"]
        ws['P%s' % idx] = "=((N%s + ($D$4 * O%s+$E$4)) / (1 -$G$5-$H$5-$I$5)) /$J$5" % (idx, idx)  # 计算最终成本价
        ws['Q%s' % idx] = "=P%s * (1 +$B$13)" % idx  # 售卖价
        ws['R%s' % idx] = "=Q%s / (1 -$D$13)" % idx  # 上货价
        idx = idx + 1

    ws = wb.get_sheet_by_name('产品链接')
    ws['B2'] = url_link
    wb.save(base_path + "03成本计算.xlsx")


if __name__ == '__main__':
    """
        实现目标：
            1、下载主图图片、详情页图片到指定目录  **处理图片大小
            2、生成成本模板并填入相关数据（尺码信息）
    """
    # 下图片
    urls = URL.split("\n")
    for url_idx in range(0, len(urls)):
        if urls[url_idx] is not None and len(urls[url_idx]) > 0:
            # 获取阿里巴巴指定路径下的商品信息
            page_info = grab_webpage(urls[url_idx])
            # 根据下载图片并根据图片类型裁剪  pil 处理
            handel_pic_info(page_info)
            # 处理库存信息
            handel_stock_inf(page_info[STOCK_INF], page_info[URL_LINK])
            if os.path.exists("C:\\Users\\Baijb\\Desktop\\产品模板\\"):
                os.rename("C:\\Users\\Baijb\\Desktop\\产品模板\\", "C:\\Users\\Baijb\\Desktop\\产品" + str(url_idx))
    # 贴标签
    # blend_brand_images(r'C:\\Users\\Baijb\\Desktop\\要上的产品\\圣诞款\\BP18\\01主图\\', r'1.jpg')
